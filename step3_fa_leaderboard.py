"""
step3_fa_leaderboard.py
=========================
Free agent CMD leaderboard -> cmd_free_agents.xlsx

Same trailing-30-day, starts-only window and genuine-starter filter as
Step 2. Qualification: 1+ start AND 5+ IP in the window, excluding anyone
rostered on any of the 12 teams. (Note: the league-wide absolute pool used
for abs_score/abs_rank still requires 2+ starts -- only the FA-only
leaderboard's inclusion and rel_score pool were loosened to 1+ start, per
explicit user request.)

Locked columns (NO ERA, NO K%/BB% -- these were removed from an earlier
draft of this file and must not come back):
  Rank, Pitcher, Starts, IP, Pure%, Ball%, CMD%, HR%, m_whip,
  rl_whip, BABIP, rel_score, abs_score, abs_rank
(this is the layout when `probables` isn't supplied -- see below for the
enriched layout, which reorders things. m_whip/rl_whip/rel_score/
abs_score/abs_rank are abbreviated headers for Model WHIP/Actual WHIP/
Score (Relative)/Score (Absolute)/Rank (Absolute), renamed per explicit
user request -- the underlying concepts and formulas are unchanged, only
the column labels.)

rel_score = min(CMD%, m_whip) percentiles within the FA-ONLY pool.
            Drives sort order and cell shading.
abs_score = same formula against the league-wide qualifying pool.
abs_rank  = literal rank position with pool size, e.g. "45 of 149".

Stoplight cell shading (NOT full-row) on Pitcher / CMD% / m_whip columns
only, using the locked hex colors below. Tiers = thirds by rel_score
rank position (top third green, middle third yellow, bottom third red).

Always writes to the SAME filename (cmd_free_agents.xlsx) so re-running the
pipeline updates the existing preview rather than creating a new file.

OPTIONAL Next Start columns (added with Step 6): when `main()` is called
with a `probables` dict (same format Step 5 uses), the xlsx gains three
extra columns -- SS (Next Start Score), Date, Opp (opponent/matchup) --
all grouped together right after Pitcher (per explicit user request),
pushing the rest of the base columns down three positions. Headers are
abbreviated and the Pitcher column shows first-initial + last name (e.g.
"M. Liberatore") to keep the sheet narrow -- see abbreviate_pitcher_name()
/ abbreviate_date_label(). Column order in this mode:
  Rank, Pitcher, SS, Date, Opp, Starts, IP, Pure%, Ball%, CMD%, HR%,
  m_whip, rl_whip, BABIP, rel_score, abs_score, abs_rank
This can't run unattended in CI (no reliable way to fetch the FanGraphs
grid without a human), so CI's default (`probables=None`) keeps writing
the original 14-column file, unreordered, unchanged. Only add these
columns from a chat session where the probables grid has already been
fetched/pasted for Step 5.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from pipeline_common import load_unified_json, percentile_rank
from step2_kondor_staff import build_pool, score_pool, LEAGUE_MIN_STARTS, LEAGUE_MIN_IP

GREEN = "9BE39B"
YELLOW = "FFEB9C"
RED = "F7A6AC"
HEADER_FILL_HEX = "2A4D3A"

# FA leaderboard inclusion is intentionally looser than the league-wide
# absolute pool (LEAGUE_MIN_STARTS=2, imported above and still used for
# abs_score/abs_rank). Per explicit user request, single-start genuine
# starters are included in the FA-only leaderboard/pool and its relative
# scoring; IP floor stays at LEAGUE_MIN_IP.
FA_MIN_STARTS = 1


def build_fa_pool(pool):
    """FA-only pool: not on any of the 12 rosters, meets FA qualification
    thresholds (looser start-count floor than the league-wide absolute pool).
    Attaches ScoreRelative computed within the FA-only pool."""
    fa_pool = {n: p for n, p in pool.items()
               if p['team_id'] is None and p['Starts'] >= FA_MIN_STARTS and p['IP'] >= LEAGUE_MIN_IP}

    cmd_vals = [p['CMD'] for p in fa_pool.values()]
    whip_vals = [p['ModelWHIP'] for p in fa_pool.values()]
    for name, d in fa_pool.items():
        cmd_pct = percentile_rank(d['CMD'], cmd_vals, higher_is_better=True)
        whip_pct = percentile_rank(d['ModelWHIP'], whip_vals, higher_is_better=False)
        d['ScoreRelative'] = min(cmd_pct, whip_pct)

    return fa_pool


def attach_next_start(fa_pool, probables, data, coefs):
    """
    Optional enrichment (chat-time only -- see module docstring on why this
    can't run unattended in CI): for each FA with at least one start in the
    `probables` dict (same format Step 5 uses -- {name: [(date_label,
    date_sort, matchup, hand), ...]}), attach that pitcher's EARLIEST
    upcoming start's date/matchup and a Start Score for that specific
    matchup (via step6_start_score.matchup_start_score, real opponent, not
    the neutral placeholder).

    Mutates fa_pool in place, adding 'NextStartDate', 'NextStartMatchup',
    'NextStartScore' (all None if the pitcher has no start in `probables`
    or the opponent abbreviation doesn't resolve). Returns fa_pool for
    chaining.

    Start Score here is percentile-ranked within the set of FA pool
    pitchers that actually have a resolvable next start -- a different
    (smaller) population than Step 5's table, so don't expect the two
    numbers to match for the same pitcher/date; they're answering slightly
    different questions (Step 5: "how does this compare among everyone
    starting that day", Step 3: "how does this compare among free agents
    with a start in the next 7 days").
    """
    from step6_start_score import build_team_offense_index, matchup_start_score
    import re

    by_team, _ = build_team_offense_index(data)

    def extract_opp_abbr(matchup):
        m = re.search(r'([A-Z]{2,4})\s*$', matchup or '')
        return m.group(1) if m else None

    predicted_whips = {}
    for name, d in fa_pool.items():
        starts = probables.get(name)
        if not starts:
            d['NextStartDate'] = None
            d['NextStartMatchup'] = None
            d['NextStartScore'] = None
            continue
        date_label, date_sort, matchup, hand = min(starts, key=lambda s: s[1])
        d['NextStartDate'] = date_label
        d['NextStartMatchup'] = matchup
        opp_abbr = extract_opp_abbr(matchup)
        predicted_whip = None
        if opp_abbr:
            m = matchup_start_score(coefs, d['CMD'] * 100, d['ModelWHIP'], opp_abbr, by_team)
            if m is not None:
                predicted_whip = m['predicted_whip']
        d['_predicted_whip_next_start'] = predicted_whip
        if predicted_whip is not None:
            predicted_whips[name] = predicted_whip

    vals = list(predicted_whips.values())
    for name, d in fa_pool.items():
        pw = d.pop('_predicted_whip_next_start', None)
        d['NextStartScore'] = percentile_rank(pw, vals, higher_is_better=False) if (pw is not None and vals) else None

    return fa_pool


def abbreviate_pitcher_name(name):
    """'Matthew Liberatore' -> 'M. Liberatore'. Keeps everything after the
    first token as-is (handles suffixes/multi-word surnames reasonably),
    just abbreviates the first name to an initial to save column width.
    Single-token names (rare, but be safe) pass through unchanged."""
    parts = (name or '').split(' ', 1)
    if len(parts) < 2 or not parts[0]:
        return name
    return f"{parts[0][0]}. {parts[1]}"


def abbreviate_date_label(date_label):
    """'Fri 8/14' -> '8/14' -- strip the leading weekday abbreviation to
    save column width; the calendar date alone is enough context here."""
    if not date_label:
        return date_label
    parts = date_label.split(' ', 1)
    return parts[1] if len(parts) == 2 else date_label


def write_xlsx(fa_pool, out_path='cmd_free_agents.xlsx', include_next_start=False):
    rows = sorted(fa_pool.values(), key=lambda d: -d['ScoreRelative'])
    n = len(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Free Agent CMD"

    # Column layout as an ordered list of (header, value_fn, number_format)
    # rather than magic column-index writes -- makes reordering (like
    # putting Next Start Score up front) a one-line change instead of
    # renumbering every ws.cell(row=r, column=N) call by hand.
    base_cols = [
        ("Rank", lambda i, d: i + 1, None),
        ("Pitcher", lambda i, d: abbreviate_pitcher_name(d['name']), None),
        ("Starts", lambda i, d: d['Starts'], None),
        ("IP", lambda i, d: round(d['IP'], 1), None),
        ("Pure%", lambda i, d: round(d['Purepct'], 4), "0.0%"),
        ("Ball%", lambda i, d: round(d['Ballpct'], 4), "0.0%"),
        ("CMD%", lambda i, d: round(d['CMD'], 4), "0.0%"),
        ("HR%", lambda i, d: round(d['HRpct'], 4), "0.0%"),
        ("m_whip", lambda i, d: round(d['ModelWHIP'], 3), "0.000"),
        ("rl_whip", lambda i, d: round(d['ActualWHIP'], 3) if d['ActualWHIP'] is not None else None, "0.000"),
        ("BABIP", lambda i, d: round(d['BABIP'], 3) if d['BABIP'] is not None else None, "0.000"),
        ("rel_score", lambda i, d: round(d['ScoreRelative'], 4), "0.00"),
        ("abs_score", lambda i, d: round(d['ScoreAbsolute'], 4) if d['ScoreAbsolute'] is not None else None, "0.00"),
        ("abs_rank", lambda i, d: d['AbsRankStr'], None),
    ]
    next_start_score_col = (
        "SS",
        lambda i, d: round(d['NextStartScore'], 4) if d.get('NextStartScore') is not None else None,
        "0.00",
    )
    next_start_extra_cols = [
        ("Date", lambda i, d: abbreviate_date_label(d.get('NextStartDate')), None),
        ("Opp", lambda i, d: d.get('NextStartMatchup'), None),
    ]

    if include_next_start:
        # Next Start Score, Next Start (date), and Next Start Matchup
        # (opponent) all sit together right after Pitcher, per explicit
        # request -- everything else keeps its relative order after that.
        cols = [base_cols[0], base_cols[1], next_start_score_col] + next_start_extra_cols + base_cols[2:]
    else:
        cols = base_cols

    headers = [c[0] for c in cols]
    ws.append(headers)

    header_fill = PatternFill(start_color=HEADER_FILL_HEX, end_color=HEADER_FILL_HEX, fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    normal_font = Font(name="Arial")
    bold_dark_font = Font(name="Arial", bold=True, color="2C2C2A")

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Stoplight shading always targets Pitcher / CMD% / m_whip by NAME,
    # not fixed column numbers -- so it stays correct regardless of layout.
    shaded_headers = {"Pitcher", "CMD%", "m_whip"}
    shaded_col_indices = [i + 1 for i, h in enumerate(headers) if h in shaded_headers]
    pitcher_col_index = headers.index("Pitcher") + 1

    for i, d in enumerate(rows):
        r = i + 2
        for col_idx, (header, value_fn, _) in enumerate(cols, start=1):
            ws.cell(row=r, column=col_idx, value=value_fn(i, d))

        frac = i / n
        color = GREEN if frac < 1 / 3 else (YELLOW if frac < 2 / 3 else RED)
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in shaded_col_indices:
            cell = ws.cell(row=r, column=col)
            cell.fill = fill
            cell.font = bold_dark_font if col == pitcher_col_index else normal_font

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            if cell.font is None or cell.font.name != "Arial":
                cell.font = normal_font

    for col_idx, (header, _, number_format) in enumerate(cols, start=1):
        if number_format:
            for r in range(2, n + 2):
                ws.cell(row=r, column=col_idx).number_format = number_format

    width_by_header = {
        "Rank": 6, "Pitcher": 14, "Starts": 8, "IP": 7, "Pure%": 8, "Ball%": 8,
        "CMD%": 8, "HR%": 7, "m_whip": 9, "rl_whip": 9, "BABIP": 8,
        "rel_score": 10, "abs_score": 10, "abs_rank": 10,
        "Date": 7, "Opp": 6, "SS": 6,
    }
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_by_header.get(header, 12)

    notes_row = n + 3
    notes = [
        "Column key: m_whip = Model WHIP, rl_whip = Actual (real) WHIP, rel_score = Score (Relative), abs_score = Score (Absolute), abs_rank = Rank (Absolute). Renamed for column width, same formulas as always.",
        "Stoplight shading (Pitcher, CMD%, m_whip) reflects rel_score: MIN of each pitcher's percentile rank on CMD% (higher=better) and m_whip (lower=better) within this free-agent-only pool -- a pitcher must be strong on BOTH to score green. Green = top third, yellow = middle third, red = bottom third, by rel_score rank position.",
        "abs_score uses the same min-of-two-percentiles formula but computed against the league-wide qualifying pool (all qualifying rostered pitchers across all 12 teams, plus free agents; 2+ starts and 5+ IP in the trailing 30-day window). abs_rank shows literal rank position within that league-wide pool.",
        "'Starts' are restricted to genuine starting pitchers: first-listed in the box score AND a season-wide average of 4.0+ IP when first-listed.",
        "Pure% and Ball% are each that pitch type's share of total pitches thrown (balls + pure strikes + in-play pitches), not batters faced. CMD% = Pure% - Ball%.",
        "K%, BB%, HR% (m_whip inputs) are correctly denominated by batters faced (BF) -- separate, correct usage from Pure%/Ball%.",
        "BABIP approximated as (H-HR)/(BF-BB-K-HR); HBP and sac flies aren't broken out in the source gamelogs.",
        "m_whip = 1.290 - 2.208*K% + 4.402*BB% + 3.766*HR%, cross-validated regression, CV R^2 = 0.58. rl_whip is real (H+BB)/IP over the same window.",
        "Free agent status computed as of the roster snapshot date plus any subsequent add/drop/trade transactions in the file.",
        "FA leaderboard inclusion floor is 1+ start (and 5+ IP) in the trailing 30-day window -- single-start genuine starters are included here. The league-wide absolute pool behind abs_score/abs_rank still requires 2+ starts, so a 1-start pitcher can appear here with a rel_score but a blank abs_score/abs_rank.",
    ]
    if include_next_start:
        notes.append(
            "SS / Date / Opp (grouped right after Pitcher): SS (Next Start Score) is Step 6's regression-fit "
            "prediction against that pitcher's ACTUAL next opponent (not a neutral placeholder), percentile-"
            "ranked within the free agents who have a resolvable next start -- a different population from "
            "rel_score/abs_score elsewhere in this sheet, so don't expect them to match. Date / Opp show "
            "the date and opponent that score was computed against, sourced from the FanGraphs probables grid "
            "available at generation time -- blank for any pitcher without a resolvable start in that window. "
            "Pitcher names are abbreviated to first initial + last name to keep columns narrow."
        )
    for j, note in enumerate(notes):
        cell = ws.cell(row=notes_row + j, column=pitcher_col_index, value=note)
        cell.font = Font(name="Arial", italic=True, size=9, color="5F5E5A")
        cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(out_path)
    return out_path, n


def main(json_path, roster_path='current_roster.json', out_path='cmd_free_agents.xlsx', probables=None):
    """
    `probables` is OPTIONAL (unlike Step 5, where it's required). When
    supplied (chat-time, same dict Step 5 uses), adds Next Start / Next
    Start Matchup / Next Start Score columns for whichever FAs have a
    start in that window. When None (CI's default, unattended run --
    there's no reliable way to fetch the FanGraphs grid unattended), the
    xlsx is written exactly as before with no schema change, so CI stays
    unaffected.
    """
    data = load_unified_json(json_path)
    with open(roster_path) as f:
        roster = json.load(f)

    pool, window_start, max_date, appearances, genuine_starters = build_pool(data, roster)
    pool, league_pool = score_pool(pool)
    fa_pool = build_fa_pool(pool)

    include_next_start = probables is not None
    if include_next_start:
        from step6_start_score import build_training_rows, fit_start_score_model
        rows, _, _, _ = build_training_rows(data)
        coefs, _ = fit_start_score_model(rows)
        attach_next_start(fa_pool, probables, data, coefs)

    path, n = write_xlsx(fa_pool, out_path, include_next_start=include_next_start)
    print(f"Wrote {path} with {n} free agents." + (" (with Next Start columns)" if include_next_start else ""))

    with open('step3_fa_pool.json', 'w') as f:
        json.dump(fa_pool, f, indent=2)

    return fa_pool


if __name__ == '__main__':
    import sys
    json_path = sys.argv[1] if len(sys.argv) > 1 else 'pennants_over_easy_unified.json'
    main(json_path)
